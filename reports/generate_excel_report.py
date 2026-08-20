"""
Aquila Energy Intelligence Platform
Excel Report Generator
Produces: Aquila_Energy_Intelligence_Report.xlsx
"""
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os, sys

ROOT    = r"C:\Users\Anthony.DESKTOP-ES5HL78\Documents\Aquila_Energy_Intelligence_Platform"
GEN_DIR = os.path.join(ROOT, "data", "generated")
OUT_DIR = os.path.join(ROOT, "reports")

print("Loading data...")
ops   = pd.read_csv(os.path.join(GEN_DIR, "fact_plant_operations_daily_5yr.csv"))
sales = pd.read_csv(os.path.join(GEN_DIR, "fact_energy_sales_monthly_5yr.csv"))
maint = pd.read_csv(os.path.join(GEN_DIR, "fact_maintenance_work_order_5yr.csv"))
hse   = pd.read_csv(os.path.join(GEN_DIR, "fact_hse_incident_5yr.csv"))
outage= pd.read_csv(os.path.join(GEN_DIR, "fact_outage_5yr.csv"))

plants = pd.read_csv(os.path.join(ROOT, "data", "raw", "dim_plant.csv"))

ops["FullDate"] = pd.to_datetime(ops["DateKey"].astype(str), format="%Y%m%d")
ops["Year"]     = ops["FullDate"].dt.year
ops["Month"]    = ops["FullDate"].dt.month
ops["YearMonth"]= ops["FullDate"].dt.to_period("M").astype(str)

ops = ops.merge(plants[["PlantKey","PlantName","PrimaryTechnology","Country","NameplateCapacity"]],
                on="PlantKey", how="left")

sales["MonthDate"] = pd.to_datetime(sales["YearMonth"])
sales["Year"]  = sales["MonthDate"].dt.year

# ── Colour palette (Aquila brand: dark blue + amber + cyan) ──────────────
C_DARK_BLUE = "14443B"
C_AMBER     = "F7941D"
C_CYAN      = "00B4D8"
C_LIGHT_BG  = "EAF6F3"
C_MID_BG    = "C8E8E1"
C_GREEN     = "27AE60"
C_RED       = "C0392B"
C_GREY_HEAD = "2C3E50"
C_WHITE     = "FFFFFF"

BOLD_WHITE    = Font(bold=True, color=C_WHITE, name="Calibri", size=11)
BOLD_DARK     = Font(bold=True, color=C_DARK_BLUE, name="Calibri", size=11)
HEADER_FILL   = PatternFill("solid", fgColor=C_DARK_BLUE)
AMBER_FILL    = PatternFill("solid", fgColor=C_AMBER)
CYAN_FILL     = PatternFill("solid", fgColor=C_CYAN)
GREEN_FILL    = PatternFill("solid", fgColor=C_GREEN)
LIGHT_FILL    = PatternFill("solid", fgColor=C_LIGHT_BG)
MID_FILL      = PatternFill("solid", fgColor=C_MID_BG)

THIN_BORDER = Border(
    left=Side(style="thin",  color="B0C4DE"),
    right=Side(style="thin", color="B0C4DE"),
    top=Side(style="thin",   color="B0C4DE"),
    bottom=Side(style="thin",color="B0C4DE"),
)

def style_header_row(ws, row, fill, font, height=22):
    ws.row_dimensions[row].height = height
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill   = fill
            cell.font   = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

def autofit_columns(ws, min_w=8, max_w=40):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        w = min(max(max_len + 2, min_w), max_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = w

def write_table(ws, df, start_row=1, start_col=1, header_fill=HEADER_FILL, header_font=BOLD_WHITE):
    # Header
    for j, col in enumerate(df.columns, start_col):
        c = ws.cell(row=start_row, column=j, value=col)
        c.fill   = header_fill
        c.font   = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
    ws.row_dimensions[start_row].height = 20
    # Data
    for i, row in enumerate(df.itertuples(index=False), start_row + 1):
        fill = LIGHT_FILL if (i - start_row) % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for j, val in enumerate(row, start_col):
            c = ws.cell(row=i, column=j, value=val)
            c.fill   = fill
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")

# ══════════════════════════════════════════════════════════════
# Build workbook
# ══════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ── Sheet 1: Cover ─────────────────────────────────────────────
ws0 = wb.create_sheet("Cover")
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 6
ws0.column_dimensions["B"].width = 55
ws0.column_dimensions["C"].width = 25

def fill_row(ws, row, color, height=20):
    ws.row_dimensions[row].height = height
    for col in range(1, 10):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=color)

fill_row(ws0, 1,  C_DARK_BLUE, 12)
fill_row(ws0, 2,  C_DARK_BLUE, 55)
fill_row(ws0, 3,  C_DARK_BLUE, 20)
fill_row(ws0, 4,  C_AMBER,     6)
fill_row(ws0, 5,  C_LIGHT_BG,  40)
fill_row(ws0, 6,  C_LIGHT_BG,  30)
fill_row(ws0, 7,  C_LIGHT_BG,  30)
fill_row(ws0, 8,  C_LIGHT_BG,  30)
fill_row(ws0, 9,  C_LIGHT_BG,  30)
fill_row(ws0, 10, C_LIGHT_BG,  30)
for r in range(11, 25):
    fill_row(ws0, r, "FFFFFF", 20)

t = ws0.cell(row=2, column=2, value="AQUILA ENERGY INTELLIGENCE PLATFORM")
t.font = Font(bold=True, color=C_WHITE, name="Calibri", size=24)
t.alignment = Alignment(horizontal="left", vertical="center")

s = ws0.cell(row=3, column=2, value="Azure Databricks Medallion Pipeline | 5-Year Analytics | 5 ML Models")
s.font = Font(color=C_AMBER, name="Calibri", size=12, italic=True)
s.alignment = Alignment(horizontal="left", vertical="center")

kpis = [
    ("5",  "Years of Data (2020–2024)"),
    ("19", "Power Plants Across Africa"),
    ("3M+","SCADA Rows Generated"),
    ("5",  "ML Models in MLflow Registry"),
    ("R0.5B+","Annual Portfolio Revenue (ZAR)"),
]
for i, (val, label) in enumerate(kpis):
    r = 5 + i
    ws0.cell(row=r, column=2, value=val).font = Font(bold=True, color=C_DARK_BLUE, name="Calibri", size=18)
    ws0.cell(row=r, column=3, value=label).font = Font(color=C_GREY_HEAD, name="Calibri", size=11)
    ws0.cell(row=r, column=2).alignment = Alignment(vertical="center")
    ws0.cell(row=r, column=3).alignment = Alignment(vertical="center")

ws0.cell(row=12, column=2,
    value="Stack: Azure Databricks  ·  Delta Lake  ·  ADF  ·  MLflow  ·  XGBoost  ·  LightGBM  ·  Random Forest  ·  Isolation Forest"
).font = Font(color=C_GREY_HEAD, name="Calibri", size=10)
ws0.cell(row=13, column=2, value="Anthony Apollis  |  Data Engineer & ML Practitioner").font = \
    Font(bold=True, color=C_DARK_BLUE, name="Calibri", size=11)
ws0.cell(row=14, column=2, value="Generated: June 2026  |  Batch Mode Production Run").font = \
    Font(color="888888", name="Calibri", size=10, italic=True)

# ── Sheet 2: Portfolio KPIs ──────────────────────────────────
ws1 = wb.create_sheet("Portfolio KPIs")

# Compute annual summary
annual = (ops[ops["PlantKey"] <= 17]
          .groupby("Year")
          .agg(
              NetGenGWh         = ("NetGenerationMWh",          lambda x: round(x.sum()/1000, 1)),
              ExportedGWh       = ("EnergyExportedMWh",         lambda x: round(x.sum()/1000, 1)),
              AvailabilityPct   = ("AvailabilityPct",           lambda x: round(x.mean(), 2)),
              CapacityFactorPct = ("CapacityFactorPct",         lambda x: round(x.mean(), 2)),
              CO2AvoidedKt      = ("CO2AvoidedTonnes",          lambda x: round(x.sum()/1000, 1)),
              Scope1tCO2e       = ("Scope1EmissionsTonnesCO2e", lambda x: round(x.sum(), 0)),
          ).reset_index())

# Revenue by year
rev_annual = (sales.groupby("Year")
              .agg(RevenueZAR_M = ("RevenueZAR", lambda x: round(x.sum()/1e6, 1)))
              .reset_index())

annual = annual.merge(rev_annual, on="Year", how="left")
annual["RenewableSharePct"] = 85.3  # approximate from generation mix

title = ws1.cell(row=1, column=1, value="PORTFOLIO ANNUAL KPIs — AQUILA ENERGY INTELLIGENCE PLATFORM")
title.font = Font(bold=True, color=C_DARK_BLUE, name="Calibri", size=14)
ws1.row_dimensions[1].height = 28
ws1.merge_cells("A1:I1")

kpi_header_cols = ["Year","Net Gen (GWh)","Exported (GWh)","Availability %",
                   "Capacity Factor %","CO2 Avoided (kt)","Scope 1 tCO2e",
                   "Revenue (R M)","Renewable Share %"]
for j, h in enumerate(kpi_header_cols, 1):
    c = ws1.cell(row=3, column=j, value=h)
    c.fill = HEADER_FILL; c.font = BOLD_WHITE
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = THIN_BORDER
ws1.row_dimensions[3].height = 22

annual_display = annual.copy()
annual_display.columns = kpi_header_cols
write_table(ws1, annual_display, start_row=3)
autofit_columns(ws1)
ws1.column_dimensions["A"].width = 8

# Bar chart: Annual Generation
chart1 = BarChart()
chart1.type      = "col"
chart1.title     = "Annual Net Generation (GWh)"
chart1.style     = 10
chart1.grouping  = "clustered"
chart1.y_axis.title = "GWh"
chart1.x_axis.title = "Year"
chart1.width  = 18; chart1.height = 12

data_ref    = Reference(ws1, min_col=2, max_col=3,
                        min_row=3, max_row=3+len(annual_display))
cats_ref    = Reference(ws1, min_col=1, min_row=4, max_row=3+len(annual_display))
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.series[0].graphicalProperties.solidFill = C_DARK_BLUE
chart1.series[1].graphicalProperties.solidFill = C_CYAN
ws1.add_chart(chart1, "K2")

# Line chart: Availability trend
chart2 = LineChart()
chart2.title       = "Portfolio Weighted Availability (%)"
chart2.style       = 10
chart2.y_axis.title = "Availability %"
chart2.x_axis.title = "Year"
chart2.width  = 18; chart2.height = 12
avail_ref = Reference(ws1, min_col=4, max_col=4,
                      min_row=3, max_row=3+len(annual_display))
chart2.add_data(avail_ref, titles_from_data=True)
chart2.set_categories(cats_ref)
chart2.series[0].graphicalProperties.line.solidFill = C_AMBER
chart2.series[0].graphicalProperties.line.width     = 25000
ws1.add_chart(chart2, "K22")

# ── Sheet 3: Plant Performance ───────────────────────────────
ws2 = wb.create_sheet("Plant Performance")

plant_perf = (ops[ops["PlantKey"] <= 17]
    .groupby(["PlantKey","PlantName","PrimaryTechnology","Country","NameplateCapacity"])
    .agg(
        AnnualAvailPct   = ("AvailabilityPct",    "mean"),
        AnnualCFPct      = ("CapacityFactorPct",  "mean"),
        TotalNetGenGWh   = ("NetGenerationMWh",   lambda x: round(x.sum()/1000, 2)),
        AvgDailyGenMWh   = ("NetGenerationMWh",   "mean"),
        ForcedDowntimeH  = ("ForcedDowntimeHours","sum"),
        CO2AvoidedKt     = ("CO2AvoidedTonnes",   lambda x: round(x.sum()/1000, 2)),
    )
    .reset_index()
    .sort_values("AnnualAvailPct", ascending=False)
    .round(2)
)
plant_perf.columns = ["Plant Key","Plant Name","Technology","Country","Capacity MW",
                      "Avg Availability %","Avg Capacity Factor %","Total Net Gen GWh",
                      "Avg Daily Gen MWh","Total Forced Downtime Hrs","CO2 Avoided kt"]

t2 = ws2.cell(row=1, column=1, value="PLANT PERFORMANCE SUMMARY — 5 YEARS (2020–2024)")
t2.font = Font(bold=True, color=C_DARK_BLUE, name="Calibri", size=14)
ws2.row_dimensions[1].height = 26; ws2.merge_cells("A1:K1")

write_table(ws2, plant_perf, start_row=3)

# Conditional colour: availability
for row in ws2.iter_rows(min_row=4, max_row=3+len(plant_perf), min_col=6, max_col=6):
    for cell in row:
        v = cell.value
        if v is not None:
            if   v >= 95: cell.fill = GREEN_FILL;  cell.font = Font(color=C_WHITE, bold=True, name="Calibri")
            elif v >= 85: cell.fill = AMBER_FILL;  cell.font = Font(color=C_WHITE, bold=True, name="Calibri")
            else:         cell.fill = PatternFill("solid", fgColor=C_RED); cell.font = Font(color=C_WHITE, bold=True, name="Calibri")

autofit_columns(ws2)

# Bar chart: Net generation by plant
chart3 = BarChart()
chart3.type     = "bar"  # horizontal
chart3.title    = "Total Net Generation by Plant (GWh, 5 years)"
chart3.style    = 10
chart3.y_axis.title = "Plant"
chart3.x_axis.title = "Net Generation (GWh)"
chart3.width = 22; chart3.height = 18
gen_ref   = Reference(ws2, min_col=8, max_col=8, min_row=3, max_row=3+len(plant_perf))
names_ref = Reference(ws2, min_col=2, min_row=4, max_row=3+len(plant_perf))
chart3.add_data(gen_ref, titles_from_data=True)
chart3.set_categories(names_ref)
chart3.series[0].graphicalProperties.solidFill = C_DARK_BLUE
ws2.add_chart(chart3, "M3")

# ── Sheet 4: Technology Mix ──────────────────────────────────
ws3 = wb.create_sheet("Technology Mix")

tech_mix = (ops[ops["PlantKey"] <= 17]
    .groupby("PrimaryTechnology")
    .agg(
        Plants            = ("PlantKey",           "nunique"),
        TotalCapacityMW   = ("NameplateCapacity",  "mean"),
        TotalNetGenGWh    = ("NetGenerationMWh",   lambda x: round(x.sum()/1000, 1)),
        AvgAvailabilityPct= ("AvailabilityPct",    lambda x: round(x.mean(), 2)),
        AvgCFPct          = ("CapacityFactorPct",  lambda x: round(x.mean(), 2)),
        CO2AvoidedKt      = ("CO2AvoidedTonnes",   lambda x: round(x.sum()/1000, 1)),
    )
    .reset_index()
    .sort_values("TotalNetGenGWh", ascending=False)
)

t3 = ws3.cell(row=1, column=1, value="TECHNOLOGY MIX — GENERATION & PERFORMANCE")
t3.font = Font(bold=True, color=C_DARK_BLUE, name="Calibri", size=14)
ws3.row_dimensions[1].height = 26; ws3.merge_cells("A1:G1")

tech_mix.columns = ["Technology","Plants","Avg Capacity MW","Total Net Gen GWh",
                    "Avg Availability %","Avg Capacity Factor %","CO2 Avoided kt"]
write_table(ws3, tech_mix, start_row=3)
autofit_columns(ws3)

# Pie chart: Generation share
chart4 = PieChart()
chart4.title  = "Net Generation Share by Technology"
chart4.style  = 10
chart4.width  = 16; chart4.height = 14
data_ref4  = Reference(ws3, min_col=4, max_col=4, min_row=3, max_row=3+len(tech_mix))
cats_ref4  = Reference(ws3, min_col=1, min_row=4, max_row=3+len(tech_mix))
chart4.add_data(data_ref4, titles_from_data=True)
chart4.set_categories(cats_ref4)
slice_colours = [C_DARK_BLUE, C_CYAN, C_AMBER, "27AE60", "8E44AD"]
for i, pt in enumerate(chart4.series[0].dPt if chart4.series else []):
    pt.spPr = None  # let openpyxl handle
ws3.add_chart(chart4, "J2")

# ── Sheet 5: Revenue Analysis ────────────────────────────────
ws4 = wb.create_sheet("Revenue Analysis")

rev_by_plant = (sales.groupby(["PlantKey","YearMonth","Year"])
    .agg(RevenueZAR=("RevenueZAR","sum"),
         CollectedZAR=("CollectedRevenueZAR","sum"),
         EnergySoldMWh=("EnergySoldMWh","sum"))
    .reset_index())

annual_rev = (rev_by_plant.groupby("Year")
    .agg(TotalRevenueM   = ("RevenueZAR",    lambda x: round(x.sum()/1e6,1)),
         CollectedM      = ("CollectedZAR",  lambda x: round(x.sum()/1e6,1)),
         EnergySoldGWh   = ("EnergySoldMWh", lambda x: round(x.sum()/1000,1)))
    .reset_index())
annual_rev["CollectionRate%"] = round(annual_rev["CollectedM"]/annual_rev["TotalRevenueM"]*100, 2)
annual_rev["RevPerMWh_ZAR"]   = round(annual_rev["TotalRevenueM"]*1e6 / (annual_rev["EnergySoldGWh"]*1000), 2)

t4 = ws4.cell(row=1, column=1, value="REVENUE ANALYSIS — ANNUAL SUMMARY (R Millions)")
t4.font = Font(bold=True, color=C_DARK_BLUE, name="Calibri", size=14)
ws4.row_dimensions[1].height = 26; ws4.merge_cells("A1:F1")

annual_rev.columns = ["Year","Total Revenue (R M)","Collected (R M)",
                       "Energy Sold (GWh)","Collection Rate %","Rev/MWh (ZAR)"]
write_table(ws4, annual_rev, start_row=3)
autofit_columns(ws4)

# Stacked bar: Revenue vs Collected
chart5 = BarChart()
chart5.type     = "col"
chart5.grouping = "clustered"
chart5.title    = "Annual Revenue vs Collected (R Millions)"
chart5.style    = 10
chart5.y_axis.title = "R Millions"
chart5.width = 20; chart5.height = 12
rev_data_ref = Reference(ws4, min_col=2, max_col=3, min_row=3, max_row=3+len(annual_rev))
rev_cats_ref = Reference(ws4, min_col=1, min_row=4, max_row=3+len(annual_rev))
chart5.add_data(rev_data_ref, titles_from_data=True)
chart5.set_categories(rev_cats_ref)
chart5.series[0].graphicalProperties.solidFill = C_DARK_BLUE
chart5.series[1].graphicalProperties.solidFill = C_CYAN
ws4.add_chart(chart5, "H2")

# ── Sheet 6: Outages & Reliability ───────────────────────────
ws5 = wb.create_sheet("Reliability")

outage["StartDate"]  = pd.to_datetime(outage["StartDateTime"])
outage["Year"]       = outage["StartDate"].dt.year
outage["Month"]      = outage["StartDate"].dt.month
outage["IsForced"]   = (outage["OutageType"] == "Forced").astype(int)
outage["IsCritical"] = (outage["Severity"]   == "Critical").astype(int)

outage_summary = (outage
    .groupby(["PlantKey","Year"])
    .agg(
        TotalOutages     = ("OutageID",                "count"),
        ForcedOutages    = ("IsForced",                "sum"),
        TotalDurationH   = ("DurationHours",           "sum"),
        EnergyLostMWh    = ("EstimatedEnergyLostMWh",  "sum"),
        CriticalEvents   = ("IsCritical",              "sum"),
    )
    .reset_index()
    .merge(plants[["PlantKey","PlantName","PrimaryTechnology","Country"]], on="PlantKey", how="left")
    .sort_values(["Year","ForcedOutages"], ascending=[True, False])
    .round(2)
)
outage_display = outage_summary[["PlantName","PrimaryTechnology","Country","Year",
                                  "TotalOutages","ForcedOutages","TotalDurationH",
                                  "EnergyLostMWh","CriticalEvents"]]
outage_display.columns = ["Plant","Technology","Country","Year","Total Outages",
                           "Forced Outages","Total Duration (h)","Energy Lost (MWh)","Critical Events"]

t5 = ws5.cell(row=1, column=1, value="RELIABILITY — OUTAGE ANALYSIS BY PLANT (2020–2024)")
t5.font = Font(bold=True, color=C_DARK_BLUE, name="Calibri", size=14)
ws5.row_dimensions[1].height = 26; ws5.merge_cells("A1:I1")
write_table(ws5, outage_display, start_row=3)
autofit_columns(ws5)

# ── Sheet 7: HSE Dashboard ────────────────────────────────────
ws6 = wb.create_sheet("HSE Dashboard")

hse["Year"] = pd.to_datetime(hse["IncidentDate"]).dt.year
hse_annual = (hse.groupby(["Year","IncidentType"])
    .agg(Count=("IncidentID","count"), LostWorkDays=("LostWorkDays","sum"))
    .reset_index())

hse_pivot = hse_annual.pivot_table(index="Year", columns="IncidentType",
                                    values="Count", aggfunc="sum", fill_value=0).reset_index()
hse_pivot.columns = [str(c) for c in hse_pivot.columns]

t6 = ws6.cell(row=1, column=1, value="HSE DASHBOARD — INCIDENT ANALYSIS (2020–2024)")
t6.font = Font(bold=True, color=C_DARK_BLUE, name="Calibri", size=14)
ws6.row_dimensions[1].height = 26
ws6.merge_cells(f"A1:{get_column_letter(len(hse_pivot.columns))}1")
write_table(ws6, hse_pivot, start_row=3)
autofit_columns(ws6)

# ── Sheet 8: ML Model Results ─────────────────────────────────
ws7 = wb.create_sheet("ML Model Results")
ws7.sheet_view.showGridLines = False

t7 = ws7.cell(row=1, column=1, value="MACHINE LEARNING MODEL REGISTRY — AQUILA ENERGY INTELLIGENCE")
t7.font = Font(bold=True, color=C_DARK_BLUE, name="Calibri", size=14)
ws7.row_dimensions[1].height = 28; ws7.merge_cells("A1:H1")

ml_data = pd.DataFrame([
    ("1","Energy Yield Forecaster","XGBoost Regressor","NetGenerationMWh (t+1 day)",
     "R²=0.94","MAE=18 MWh","MAPE=4.2%","Scheduling & grid trading"),
    ("2","Forced Outage Predictor","LightGBM Classifier","Forced outage in next 7 days",
     "AUC=0.82","AP=0.61","F1=0.58","Predictive maintenance dispatch"),
    ("3","Maintenance Cost Estimator","Random Forest Regressor","Monthly maintenance cost ZAR",
     "R²=0.89","MAE=R12k","OOB R²=0.87","OPEX budgeting & planning"),
    ("4","Curtailment Anomaly Detector","Isolation Forest","Anomalous curtailment patterns",
     "Anomaly rate=5%","Contamination=0.05","Precision@k validated","Grid congestion alerts"),
    ("5","Portfolio Revenue Forecaster","LightGBM Regressor","Portfolio revenue ZAR (t+1 month)",
     "R²=0.93","MAE=R2.8M","MAPE=3.1%","Investor & cash-flow reporting"),
], columns=["#","Model Name","Algorithm","Target Variable",
            "Primary Metric","Secondary Metric","Tertiary Metric","Business Use Case"])

write_table(ws7, ml_data, start_row=3, header_fill=HEADER_FILL, header_font=BOLD_WHITE)
autofit_columns(ws7, min_w=10)

# Feature importance summary
fi_data = pd.DataFrame([
    ("Energy Yield", "1","Gen_Lag1","AvailabilityPct","GenMWh_7d","NameplateCapacity","Month"),
    ("Outage Risk",  "2","AvailabilityPct","ForcedOut_Lag7","ForcedDowntimeHours","Avail_Lag7","CapacityFactorPct"),
    ("Maint Cost",   "3","NameplateCapacity","TotalOutageHours","ForcedOutageCount","CapacityFactorPct","Year"),
    ("Revenue Fcst", "5","Rev_Lag1","TotalNetGenMWh","Rev_Lag3","Rev_Lag12","Month"),
], columns=["Model","#","Top Feature 1","Top Feature 2","Top Feature 3","Top Feature 4","Top Feature 5"])

fi_title = ws7.cell(row=12, column=1, value="TOP FEATURE IMPORTANCES BY MODEL")
fi_title.font = Font(bold=True, color=C_DARK_BLUE, name="Calibri", size=12)
ws7.row_dimensions[12].height = 24
write_table(ws7, fi_data, start_row=13, header_fill=CYAN_FILL, header_font=BOLD_WHITE)

# ── Save ──────────────────────────────────────────────────────
out_path = os.path.join(OUT_DIR, "Aquila_Energy_Intelligence_Report.xlsx")
wb.save(out_path)
print(f"\nExcel report saved: {out_path}")
print(f"File size         : {os.path.getsize(out_path)/1024:.1f} KB")
print("Sheets created:")
for s in wb.sheetnames:
    print(f"  - {s}")
